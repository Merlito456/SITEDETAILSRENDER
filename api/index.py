from flask import Flask, request, jsonify
from flask_cors import CORS
import hashlib
import json
import os
from datetime import datetime, timedelta
import urllib.parse
import requests
import sys
import tempfile
import traceback

app = Flask(__name__)
CORS(app)

SECRET_KEY = os.environ.get('SECRET_KEY', "YOUR_SECRET_KEY_HERE_CHANGE_THIS_TO_A_RANDOM_STRING_12345")

# Global error handler
@app.errorhandler(Exception)
def handle_exception(e):
    return jsonify({
        'error': str(e),
        'traceback': traceback.format_exc(),
        'type': type(e).__name__
    }), 500

def load_excel_data(file_path):
    """Load Excel data using openpyxl directly without pandas"""
    try:
        import openpyxl
        if not os.path.exists(file_path):
            return None
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        if not headers:
            return None
        
        # Convert to list of dicts
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and any(cell is not None for cell in row):
                row_dict = {}
                for idx, value in enumerate(row):
                    if idx < len(headers):
                        row_dict[headers[idx]] = value
                data.append(row_dict)
        
        print(f"✅ Loaded {len(data)} records from {file_path}")
        return data
    except Exception as e:
        print(f"❌ Error loading {file_path}: {str(e)}")
        return None

def get_site_by_plaid(data, plaid):
    """Find site by PLAID in the data list"""
    if not data:
        return None
    
    plaid = str(plaid).strip()
    for record in data:
        if record.get('PLAID') and str(record.get('PLAID')).strip() == plaid:
            return record
    return None

def get_online_time():
    try:
        time_apis = [
            "https://worldtimeapi.org/api/timezone/Etc/UTC",
            "https://timeapi.io/api/time/current/utc",
        ]
        for api in time_apis:
            try:
                response = requests.get(api, timeout=5)
                if response.status_code == 200:
                    data = response.json()
                    if 'utc_datetime' in data:
                        return datetime.fromisoformat(data['utc_datetime'].replace('Z', '+00:00'))
                    elif 'dateTime' in data:
                        return datetime.fromisoformat(data['dateTime'].replace('Z', '+00:00'))
                    elif 'unixtime' in data:
                        return datetime.fromtimestamp(data['unixtime'])
            except:
                continue
        return None
    except:
        return None

def validate_device(device_to_check, allowed_devices_hashed):
    if not device_to_check or not allowed_devices_hashed:
        return False
    
    hashed = hashlib.sha256(device_to_check.encode()).hexdigest()
    if hashed in allowed_devices_hashed:
        return True
    
    if device_to_check.startswith('MAC:'):
        without_prefix = device_to_check[4:]
        hashed = hashlib.sha256(without_prefix.encode()).hexdigest()
        if hashed in allowed_devices_hashed:
            return True
    elif device_to_check.startswith('IMEI:'):
        without_prefix = device_to_check[5:]
        hashed = hashlib.sha256(without_prefix.encode()).hexdigest()
        if hashed in allowed_devices_hashed:
            return True
    
    return False

def validate_token(token, data, device_fingerprint=None):
    try:
        token = token.strip()
        if '%' in token:
            token = urllib.parse.unquote(token)
        
        try:
            token_data = bytes.fromhex(token).decode('utf-8')
        except ValueError:
            return None, "Invalid token format - not valid hex"
        
        parts = token_data.rsplit('|', 1)
        if len(parts) != 2:
            return None, "Invalid token format - expected 2 parts"
        
        payload_json, signature = parts
        expected_signature = hashlib.sha256(f"{payload_json}|{SECRET_KEY}".encode()).hexdigest()
        if signature != expected_signature:
            return None, "Invalid token signature - token may be tampered"
        
        try:
            payload = json.loads(payload_json)
        except json.JSONDecodeError:
            return None, "Invalid token payload"
        
        created_str = payload.get('c')
        expires_str = payload.get('e')
        site_plaid = payload.get('s')
        user_name = payload.get('u')
        user_email = payload.get('a')
        allowed_devices = payload.get('d', [])
        raw_devices = payload.get('raw', '')
        
        if not all([created_str, expires_str, site_plaid]):
            return None, "Missing required token data"
        
        current_time = get_online_time()
        if current_time is None:
            current_time = datetime.utcnow()
        
        try:
            created = datetime.fromisoformat(created_str)
            expires = datetime.fromisoformat(expires_str)
        except ValueError:
            return None, "Invalid date format"
        
        if current_time > expires:
            return None, f"Token expired on {expires.strftime('%B %d, %Y')}"
        
        if current_time < created - timedelta(minutes=5):
            return None, "Token is from the future - possible fraud"
        
        if allowed_devices and device_fingerprint:
            if not validate_device(device_fingerprint, allowed_devices):
                return None, "Device not authorized. MAC Address or IMEI not recognized."
        elif allowed_devices:
            return None, "Device verification required. Please ensure your device is registered."
        
        site_data = get_site_by_plaid(data, site_plaid)
        if site_data is None:
            return None, f"Site not found: {site_plaid}"
        
        site_data['_user_email'] = user_email
        site_data['_user_name'] = user_name
        site_data['_token_created'] = created_str
        site_data['_token_expires'] = expires_str
        site_data['_device_restricted'] = bool(allowed_devices)
        site_data['_device_count'] = len(allowed_devices)
        site_data['_raw_devices'] = raw_devices
        
        return site_data, None
        
    except Exception as e:
        return None, f"Validation error: {str(e)}"

@app.route('/', methods=['GET'])
def home():
    return jsonify({
        'name': 'GPS Extractor API (Vercel)',
        'version': '1.0',
        'status': 'running',
        'platform': 'Vercel',
        'endpoints': {
            '/': 'This info page',
            '/api/health': 'Health check',
            '/api/validate': 'Validate token with device fingerprint'
        }
    })

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint - returns API status"""
    return jsonify({
        'status': 'healthy',
        'platform': 'Vercel',
        'message': 'Flask API is running on Vercel',
        'python_version': sys.version,
        'endpoints': {
            '/': 'API Info',
            '/api/health': 'Health check',
            '/api/validate': 'Token validation (GET/POST)'
        }
    })

@app.route('/api/validate', methods=['GET', 'POST'])
def api_validate():
    try:
        if request.method == 'POST':
            data = request.get_json()
            if data:
                token = data.get('token', '')
                device_fingerprint = data.get('device_fingerprint', '')
            else:
                token = request.form.get('token', '')
                device_fingerprint = request.form.get('device_fingerprint', '')
        else:
            token = request.args.get('token', '')
            device_fingerprint = request.args.get('device_fp', '')
        
        if not token:
            return jsonify({
                'success': False,
                'error': 'Missing token parameter'
            })
        
        # Try multiple paths for database file
        data = None
        possible_paths = [
            "/tmp/database.xlsx",
            "database.xlsx",
            "data/database.xlsx",
            "/opt/render/project/src/database.xlsx",
            "/opt/render/project/src/data/database.xlsx"
        ]
        
        for path in possible_paths:
            data = load_excel_data(path)
            if data is not None:
                break
        
        # If no local file, try to download from GitHub
        if data is None:
            try:
                github_url = "https://raw.githubusercontent.com/Merlito456/SITEDETAILSRENDER/main/database.xlsx"
                response = requests.get(github_url, timeout=10)
                if response.status_code == 200:
                    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                    temp_file.write(response.content)
                    temp_file.close()
                    data = load_excel_data(temp_file.name)
            except Exception as e:
                print(f"Error downloading database: {e}")
        
        if data is None:
            return jsonify({
                'success': False,
                'error': 'No data available. Please upload database.xlsx'
            })
        
        site_data, error = validate_token(token, data, device_fingerprint)
        
        if site_data:
            return jsonify({
                'success': True,
                'data': site_data
            })
        else:
            return jsonify({
                'success': False,
                'error': error or 'Validation failed'
            })
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Server error: {str(e)}'
        })

# For Vercel
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print(f"🚀 Starting Flask API on port {port}")
    app.run(host='0.0.0.0', port=port, debug=False)
